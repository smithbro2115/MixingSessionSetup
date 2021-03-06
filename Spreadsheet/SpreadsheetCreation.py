import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from shutil import copyfile
from Functionality.useful_utils import resource_path


def create_new_spreadsheet_from_template(path, template_path):
    copyfile(template_path, path)


def create_new_sfx_list_spreadsheet(path):
    create_new_spreadsheet_from_template(path, resource_path("SFX LIST TEMPLATE.xlsx"))


def set_scene_rule(sheet):
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(fill=red_fill)
    r = Rule(type="expression", dxf=dxf)
    r.formula = ['=AND(LEFT($C2, 6)<>"SOUND:", $C2<>"")']
    sheet.conditional_formatting.add("A2:Y1000", r)


def populate_spreadsheet(spreadsheet_path, scenes_and_sounds):
    spreadsheet = openpyxl.load_workbook(spreadsheet_path)
    sheet = spreadsheet.active
    scene_number = 1
    for i, val in enumerate(scenes_and_sounds):
        row_num = i+2
        if val['type'] == "scene_header":
            scene_number = val['scene_number']
        sound_name_format = sheet.cell(row_num, 3).style
        sheet.cell(row_num, 1).value = scene_number
        sheet.cell(row_num, 2).value = row_num-1
        sheet.cell(row_num, 3).value = val["content"]
        sheet.cell(row_num, 3).style = sound_name_format
    # set_scene_rule(sheet)
    spreadsheet.save(spreadsheet_path)
